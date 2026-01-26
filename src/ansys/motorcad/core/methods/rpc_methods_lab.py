# Copyright (C) 2022 - 2026 ANSYS, Inc. and/or its affiliates.
# SPDX-License-Identifier: MIT
#
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""RPC methods for Motor-CAD Lab."""

k_num_custom_losses_internal_lab = "NumCustomLossesInternal_Lab"
k_custom_loss_name_internal_lab = "CustomLoss_name_internal_lab"
k_custom_loss_function_internal_lab = "CustomLoss_Function_Internal_Lab"
k_custom_loss_type_internal_lab = "CustomLoss_Type_Internal_Lab"
k_custom_loss_thermal_node_internal_lab = "CustomLoss_ThermalNode_Internal_Lab"
k_num_custom_losses_external_lab = "NumCustomLossesExternal_Lab"
k_custom_loss_name_external_lab = "CustomLoss_Name_External_Lab"
k_custom_loss_power_function_external_lab = "CustomLoss_PowerFunction_External_Lab"
k_custom_loss_voltage_function_external_lab = "CustomLoss_VoltageFunction_External_Lab"


from openpyxl import Workbook

try:
    import numpy as np
    from scipy.io import loadmat

    Num_Sci_py_AVAILABLE = True
except ImportError:
    Num_Sci_py_AVAILABLE = False


def _write_excel(data, sheets, DC_voltage_list, i, wb, reset_speeds):
    i_len, j_len = data["Speed"].shape
    for sheet in sheets:
        ws = wb.create_sheet("Newsheet")
        ws.title = "Voltages"
        if len(DC_voltage_list) > 1:
            ws.title = sheet + str(i + 1)
        else:
            ws.title = sheet
        for jj, col in enumerate(ws.iter_cols(min_col=0, max_col=j_len, max_row=i_len)):
            for ii, cell in enumerate(col):
                if reset_speeds and sheet == "Speed" and data[sheet][ii][jj] == 1:
                    # Special case for induction motor, where we want to
                    # change speed from 1 RPM to 0 RPM
                    ws[cell.coordinate] = 0
                else:
                    ws[cell.coordinate] = data[sheet][ii][jj]
    return wb


class _RpcMethodsLab:
    def __init__(self, mc_connection):
        self.connection = mc_connection

    def calculate_test_performance_lab(self):
        """Calculate the test performance.

        Results are saved in the MOT file results folder as ``MotorLAB_caldata.mat``.
        """
        method = "CalculateTestPerformance_Lab"
        return self.connection.send_and_receive(method)

    def export_duty_cycle_lab(self):
        """Export the calculated duty cycle data to the thermal model."""
        method = "ExportDutyCycle_Lab"
        return self.connection.send_and_receive(method)

    def get_model_built_lab(self):
        """Test if the Lab model must be built or rebuilt before running calculations.

        Returns
        -------
        bool
            ``True`` if the Lab model has been built and is valid for the current settings,
            ``False`` otherwise.
        """
        method = "GetModelBuilt_Lab"
        return self.connection.send_and_receive(method)

    def show_results_viewer_lab(self, calculation_type):
        """Load the results viewer for the specified Lab calculation type.

        Parameters
        ----------
        calculation_type : str
            Type of calculation. Options are ``"Electromagnetic"``, ``"Thermal"``,
            ``"Generator"``, ``"Duty Cycle"``, and ``"Calibration"``.
        """
        method = "ShowResultsViewer_Lab"
        params = [calculation_type]
        return self.connection.send_and_receive(method, params)

    def export_figure_lab(self, calculation_type, variable, file_name):
        """Export an image of the Lab results graph.

        Parameters
        ----------
        calculation_type : str
            Type of calculation. Options are ``"Electromagnetic"``, ``"Thermal"``,
            ``"Generator"``, ``"Duty Cycle"``, and ``"Calibration"``.
        variable : str
           Variable to plot on the Y axis (2D graphs) or Z axis (3D graphs). For
           example, ``"Shaft Torque"``.
        file_name : str
            Name of the image file.
        """
        method = "ExportFigure_Lab"
        params = [calculation_type, variable, file_name]
        return self.connection.send_and_receive(method, params)

    def calculate_generator_lab(self):
        """Calculate generator performance.

        Results are saved in the MOT file results folder as ``LabResults_Generator.mat``.
        """
        method = "CalculateGenerator_Lab"
        return self.connection.send_and_receive(method)

    def load_external_model_lab(self, file_path):
        """Load an external model data file.

        This parameter is used when the Lab link type is set to ``Custom`` or ``Ansys Maxwell``.

        Parameters
        ----------
        file_path : str
            Full path to the data file, including the file name. Use the ``r'filepath'``
            syntax to force Python to ignore special characters.
        """
        method = "LoadExternalModel_Lab"
        params = [file_path]
        return self.connection.send_and_receive(method, params)

    def clear_model_build_lab(self):
        """Clear the Lab model build."""
        method = "ClearModelBuild_Lab"
        return self.connection.send_and_receive(method)

    def build_model_lab(self):
        """Build the Lab model."""
        method = "BuildModel_Lab"
        return self.connection.send_and_receive(method)

    def calculate_operating_point_lab(self):
        """Run the Lab operating point calculation."""
        method = "CalculateOperatingPoint_Lab"
        return self.connection.send_and_receive(method)

    def calculate_magnetic_lab(self):
        """Run the Lab magnetic calculation."""
        method = "CalculateMagnetic_Lab"
        return self.connection.send_and_receive(method)

    def calculate_thermal_lab(self):
        """Run the Lab thermal calculation."""
        method = "CalculateThermal_Lab"
        return self.connection.send_and_receive(method)

    def calculate_duty_cycle_lab(self):
        """Run the Lab duty cycle."""
        method = "CalculateDutyCycle_Lab"
        return self.connection.send_and_receive(method)

    def add_internal_custom_loss(self, name, function, type, thermal_node):
        """Add an internal custom loss.

        Parameters
        ----------
        name : str
            Name of lab internal custom loss
        function : str
            Function of lab internal custom loss
        type : str
            Type of lab internal custom loss. Options are ``Electrical`` or ``Mechanical``
        thermal_node : int
            Thermal node of lab internal custom loss

        """
        type = type.capitalize()
        # Internal Custom Loss Type is case-sensitive in MotorCAD.
        # Added a line to match the required format.
        if type not in ["Electrical", "Mechanical"]:
            raise ValueError("Thermal Loss Type must be Electrical or Mechanical")
        if not self.get_node_exists(thermal_node):
            raise ValueError("Thermal node does not exist")
        else:
            no_internal_losses = self.get_variable(k_num_custom_losses_internal_lab)
            self.set_variable(k_num_custom_losses_internal_lab, no_internal_losses + 1)
            self.set_array_variable(k_custom_loss_name_internal_lab, no_internal_losses, name)
            self.set_array_variable(
                k_custom_loss_function_internal_lab, no_internal_losses, function
            )
            self.set_array_variable(k_custom_loss_type_internal_lab, no_internal_losses, type)
            self.set_array_variable(
                k_custom_loss_thermal_node_internal_lab, no_internal_losses, thermal_node
            )

    def add_external_custom_loss(self, name, power_function, voltage_function):
        """Add an external custom loss.

        Parameters
        ----------
        name : str
            Name of lab external custom loss
        power_function : str
            Power function for lab external custom loss
        voltage_function : str
            Function for voltage drop for lab external custom loss

        """
        no_external_losses = self.get_variable(k_num_custom_losses_external_lab)
        self.set_variable(k_num_custom_losses_external_lab, no_external_losses + 1)
        self.set_array_variable(k_custom_loss_name_external_lab, no_external_losses, name)
        self.set_array_variable(
            k_custom_loss_power_function_external_lab, no_external_losses, power_function
        )
        self.set_array_variable(
            k_custom_loss_voltage_function_external_lab, no_external_losses, voltage_function
        )

    def remove_internal_custom_loss(self, name):
        """Remove an internal custom loss by name.

        Parameters
        ----------
        name : str
            Name of lab internal custom loss

        """
        index = self._get_index_from_name(
            name, k_num_custom_losses_internal_lab, k_custom_loss_name_internal_lab
        )
        self._motorcad_array_pop(
            index,
            k_num_custom_losses_internal_lab,
            [
                k_custom_loss_name_internal_lab,
                k_custom_loss_function_internal_lab,
                k_custom_loss_type_internal_lab,
                k_custom_loss_thermal_node_internal_lab,
            ],
        )

    def remove_external_custom_loss(self, name):
        """Remove an external custom loss by name.

        Parameters
        ----------
        name : str
            Name of lab external custom loss

        """
        index = self._get_index_from_name(
            name, k_num_custom_losses_external_lab, k_custom_loss_name_external_lab
        )
        self._motorcad_array_pop(
            index,
            k_num_custom_losses_external_lab,
            [
                k_custom_loss_name_external_lab,
                k_custom_loss_power_function_external_lab,
                k_custom_loss_voltage_function_external_lab,
            ],
        )

    def _motorcad_array_pop(self, index, var_length_array, list_of_var_names):
        """Remove variables at a specified location within an array.

        Parameters
        ----------
        index : int
            Index of array to remove
        var_length_array : str
            Variable which specifies the length of the array
        list_of_var_names : list
            List of variables to pop

        """
        array_length = self.get_variable(var_length_array)

        for i in range(index + 1, array_length):
            for j in range(len(list_of_var_names)):
                self.set_array_variable(
                    list_of_var_names[j], i - 1, self.get_array_variable(list_of_var_names[j], i)
                )

        self.set_variable(var_length_array, array_length - 1)

    def _get_index_from_name(self, name, var_length_array, variable_name):
        """Retrieve index of a specified variable name within an array.

        Parameters
        ----------
        name : str
            Specified name to find
        var_length_array : str
            Variable which specifies the length of the array
        variable_name : str
            Variable that holds the list of names

        """
        array_length = self.get_variable(var_length_array)
        for i in range(array_length):
            if name == self.get_array_variable(variable_name, i):
                index = i
                break
        else:
            raise NameError("Provided name is not listed")
        return index

    def export_lab_model(self, file_path):
        """Export lab model.

        Parameters
        ----------
        file_path : str
            File path including lab model file name and file extension (.lab)
        """
        method = "ExportLabModel"
        params = [file_path]
        return self.connection.send_and_receive(method, params)

    def _set_model_parameters(self, **kwargs):
        if "Max_speed" in kwargs:
            self.set_variable("SpeedMax_MotorLAB", kwargs["Max_speed"])
        if "Min_speed" in kwargs:
            if self.get_variable("Motor_Type") == 1 and kwargs["Min_speed"] == 0:
                self.set_variable("SpeedMin_MotorLAB", 1)
            else:
                self.set_variable("SpeedMin_MotorLAB", kwargs["Min_speed"])
        if "Speed_step" in kwargs:
            self.set_variable("Speedinc_MotorLAB", kwargs["Speed_step"])

        Current_def = self.get_variable("CurrentSpec_MotorLAB")
        if Current_def == 0:  # peak
            if "I_max" in kwargs:
                if self.get_variable("Motor_Type") == 6:  # Sync
                    self.set_variable("Sync_StatorCurrentMax_Lab", kwargs["I_max"])
                else:
                    self.set_variable("Imax_MotorLAB", kwargs["I_max"])
            if "I_min" in kwargs:
                self.set_variable("Imin_MotorLAB", kwargs["I_min"])
        else:  # RMS
            if "I_max" in kwargs:
                if self.get_variable("Motor_Type") == 6:  # Sync
                    self.set_variable("Sync_StatorCurrentMax_RMS_Lab", kwargs["I_max"])
                else:
                    self.set_variable("Imax_RMS_MotorLAB", kwargs["I_max"])
            if "I_min" in kwargs:
                self.set_variable("Imin_RMS_MotorLAB", kwargs["I_min"])

        if "I_inc" in kwargs:
            if self.get_variable("Motor_Type") == 6:  # Sync machine
                print("sync executed")
                self.set_variable("Sync_CurrentIncs_Lab", kwargs["I_inc"])
            else:
                self.set_variable("Iinc_MotorLAB", kwargs["I_inc"])

        # choose motoring, generating or both modes
        if "Rotor_current_max" in kwargs:
            self.set_variable("Sync_RotorCurrentMax_Lab", kwargs["Rotor_current_max"])

        if "Op_mode" in kwargs:
            self.set_variable("OperatingMode_Lab", kwargs["Op_mode"])

    def export_concept_ev_model(self, **kwargs):
        """Export efficiency map in concept ev excel format.

        This will run an efficiency map calculation in Motor-CAD Lab and export the
        results for use in ConceptEV. Lab variables such as maximum and minimum
        speed and current will be changed by this method.

        Parameters
        ----------
        Max_speed : int
            Maximum speed in electromagnetic calculation
        Min_speed : int
            Minimum speed in electromagnetic calculation
        Speed_step : int
            Speed increment in electromagnetic calculation
        I_max : float
            Maximum current (peak or rms based on settings)
        I_min : float
            Minimum current (peak or rms based on settings)
        I_inc : float
            Current increment in electromagnetic calculation
        Rotor_current_max: float
            Maximum rotor current in electromagnetic calculation (only in Sync machines)
        Op_mode: int
            0 Motor, 1 Generator, 2 Motor / Generator mode
        DC_voltage_list: list
            List of DC bus voltages
        """
        if not Num_Sci_py_AVAILABLE:
            raise ImportError(
                "Failed to export concept_ev_model. Please ensure Numpy and Scipy are installed"
            )

        save_message_display_state = self.get_variable("MessageDisplayState")
        try:
            self.set_variable("MessageDisplayState", 2)
            self.set_motorlab_context()
            file_path = self.get_variable("ResultsPath_MotorLAB") + "ConceptEV_elecdata.xlsx"
            # set model parameters
            _RpcMethodsLab._set_model_parameters(self, **kwargs)
            wb = Workbook()
            # choose number of DC bus voltages (list as user input)
            if "DC_voltage_list" in kwargs:
                DC_voltage_list = kwargs["DC_voltage_list"]
            else:
                DC_voltage_list = [self.get_variable("DCBusVoltage")]

            ws = wb.active
            ws.title = "Voltages"
            ws["A1"] = "Index"
            ws["B1"] = "Voltages"
            for i, DC_voltage in enumerate(DC_voltage_list):
                ws["A" + str(i + 2)] = i + 1
                ws["B" + str(i + 2)] = DC_voltage_list[i]

            # Units sheet

            # set _calcualtion type  Efficiency Map
            self.set_variable("EmagneticCalcType_Lab", 1)
            sheets = [
                "Speed",
                "Shaft_Torque",
                "Stator_Current_Line_RMS",
                "Total_Loss",
                "Power_Factor",
            ]
            for i, DC_voltage in enumerate(DC_voltage_list):
                self.set_variable("DCBusVoltage", DC_voltage)
                # run Efficiency Map calculation
                self.calculate_magnetic_lab()
                # read the lab data .mat file
                data_file_path = self.get_variable("ResultsPath_MotorLAB") + "MotorLAB_elecdata.mat"
                data = loadmat(data_file_path)
                if (
                    "Min_speed" in kwargs
                    and self.get_variable("Motor_Type") == 1
                    and kwargs["Min_speed"] == 0
                ):
                    wb = _write_excel(data, sheets, DC_voltage_list, i, wb, True)
                else:
                    wb = _write_excel(data, sheets, DC_voltage_list, i, wb, False)

            units = [
                "Power_Factor",
                "Total_Loss",
                "Stator_Current_Line_RMS",
                "Shaft_Torque",
                "Speed",
            ]
            ws = wb.create_sheet("Newsheet")
            ws.title = "Units"
            for i, unit in enumerate(units):
                ws["A" + str(i + 1)] = unit
                index = np.where(np.strings.find(data["varStr"], unit) == 0)[0]
                ws["B" + str(i + 1)] = data["varUnits"][index][0]
            wb.save(file_path)
        finally:
            self.set_variable("MessageDisplayState", save_message_display_state)
