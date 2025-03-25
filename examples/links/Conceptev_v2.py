#%%
from scipy.io import loadmat
from openpyxl import Workbook
import ansys.motorcad.core as pymotorcad
import numpy as np

mc = pymotorcad.MotorCAD(keep_instance_open=True)
work_dir="C:\Consusltancy\Working\Concept_ev_maps"
mot_file_name="e5_eMobility_IM_150kW_User" #""e9_eMobility_IPM_User"
mot_file=work_dir+"\\"+mot_file_name+".mot"
mc.load_from_file(mot_file)
mc.set_variable("MessageDisplayState", 2)
mc.set_motorlab_context()
Excel_location=work_dir+"\\"+"Map_exportIM.xlsx"
#%%-----------------------------------------
# function takes following  parameters as keword arguments 
# User inputs, can be provided by user or by default existing mot file values will be taken 
# set the speed range
# Max_speed
# Speed_Step
# Min_speed
# set the current range
# I_max
# I_min
# I_inc
# Op_mode=0 # 0= Motor, 1= Generator , 2 = Motor+Generator
# DC_voltage_list=[400,600]
def write_excel_IM_speed(data,sheets,DC_voltage_list,i,wb):
    i_len, j_len = data["Speed"].shape
    for sheet in sheets:
        ws=wb.create_sheet("Newsheet")
        ws.title="Voltages"
        if len(DC_voltage_list)>1:
            ws.title=sheet+str(i+1)
        else:
            ws.title=sheet
        ii=0
        jj=0
        if sheet=="Speed":
            for col in ws.iter_cols(min_col=0,max_col=j_len,max_row=i_len):
                ii=0
                for cell in col:
                    
                    if data[sheet][ii][jj]==1:
                        ws[cell.coordinate]=0
                    else:
                        ws[cell.coordinate]=data[sheet][ii][jj]
                    ii=ii+1
                jj=jj+1
        else:
            for col in ws.iter_cols(min_col=0,max_col=j_len,max_row=i_len):
                ii=0
                for cell in col:
                    ws[cell.coordinate]=data[sheet][ii][jj]
                    ii=ii+1
                jj=jj+1
    return wb


def write_excel_BPM(data,sheets,DC_voltage_list,i,wb):
    i_len, j_len = data["Speed"].shape
    for sheet in sheets:
        ws=wb.create_sheet("Newsheet")
        ws.title="Voltages"
        if len(DC_voltage_list)>1:
            ws.title=sheet+str(i+1)
        else:
            ws.title=sheet
        ii=0
        jj=0
        for col in ws.iter_cols(min_col=0,max_col=j_len,max_row=i_len):
            ii=0
            for cell in col:
                ws[cell.coordinate]=data[sheet][ii][jj]
                ii=ii+1
            jj=jj+1
    return wb

def concept_ev_exportexcel(mc, work_dir,mot_file_name,Excel_location, **kwargs ):


    if "Max_speed" in kwargs:
        mc.set_variable("SpeedMax_MotorLAB",kwargs["Max_speed"])
    if "Min_speed" in kwargs:
        if mc.get_variable("Motor_Type")==1 and kwargs["Min_speed"]==0:
            mc.set_variable("SpeedMin_MotorLAB",1)
        else:
            mc.set_variable("SpeedMin_MotorLAB",kwargs["Min_speed"])

    if "Speed_Step" in kwargs:
        mc.set_variable("Speedinc_MotorLAB",kwargs["Speed_Step"])


    # set current range
    # check weather a RMS or Peak current def
    Current_def=mc.get_variable("CurrentSpec_MotorLAB")

    if Current_def==0: # peak 
        if "I_max" in kwargs:
            if mc.get_variable("Motor_Type")==5:#Sync
                mc.set_variable("Sync_StatorCurrentMax_Lab",kwargs["I_max"])
            else:
                mc.set_variable("Imax_MotorLAB",kwargs["I_max"])
        if "I_min" in kwargs:
            mc.set_variable("Imin_MotorLAB",kwargs["I_min"])
        if "I_inc" in kwargs:
            mc.set_variable("Iinc_MotorLAB",kwargs["I_inc"])
    else: #RMS
        if "I_max" in kwargs:
                if mc.get_variable("Motor_Type")==5: #Sync
                    mc.set_variable("Sync_StatorCurrentMax_RMS_Lab",kwargs["I_max"])
                else: 
                    mc.set_variable("Imax_RMS_MotorLAB",kwargs["I_max"])
        if "I_min" in kwargs:
            mc.set_variable("Imin_RMS_MotorLAB",kwargs["I_min"])
        if "I_inc" in kwargs:
            if mc.get_variable("Motor_Type")==5: #Sync machine
                mc.set_variable("Sync_CurrentIncs_Lab",kwargs["I_inc"])
            else:
                mc.set_variable("Iinc_MotorLAB",kwargs["I_inc"])

    # choose motoring, generating or both modes
    if "Rotor_current_max" in kwargs:
        mc.set_variable("Sync_RotorCurrentMax_Lab",kwargs["Rotor_current_max"])

    if "Op_mode" in kwargs:
        mc.set_variable("OperatingMode_Lab",kwargs["Op_mode"])

    wb = Workbook()
    # choose number of DC bus voltages (list as user input)

    if "DC_voltage_list" in kwargs:
        DC_voltage_list= kwargs["DC_voltage_list"]
    else:
        DC_voltage_list= [mc.get_variable("DCBusVoltage")]
  
    ws=wb.active
    ws.title="Voltages"
    ws['A1']="Index"
    ws['B1']="Voltages"
    for i,DC_voltage in enumerate(DC_voltage_list):
        ws['A'+str(i+2)]=i+1
        ws['B'+str(i+2)]=DC_voltage_list[i]

    # Units sheet
    
    # set _calcualtion type  Efficiency Map
    mc.set_variable("EmagneticCalcType_Lab",1)
    sheets=["Speed", "Shaft_Torque","Stator_Current_Line_RMS","Total_Loss","Power_Factor"]
    for i,DC_voltage in enumerate(DC_voltage_list):
        mc.set_variable("DCBusVoltage",DC_voltage)
    # run efficiency map calculation 
        mc.calculate_magnetic_lab()
    # read the efficieny map .mat file 
        data_file_path=work_dir+"\\"+mot_file_name+"\\Lab\\MotorLAB_elecdata.mat"
        data=loadmat(data_file_path)
        if mc.get_variable("Motor_Type")==1 and kwargs["Min_speed"]==0:
            wb= write_excel_IM_speed(data,sheets,DC_voltage_list,i,wb)
        else:
            wb=write_excel_BPM(data,sheets,DC_voltage_list,i,wb )


   
    units=["Power_Factor","Total_Loss","Stator_Current_Line_RMS","Shaft_Torque","Speed"]
    ws=wb.create_sheet("Newsheet")
    ws.title="Units"
    for i,unit in enumerate(units):
        ws['A'+str(i+1)]=unit
        index=np.where(np.strings.find(data["varStr"],unit)==0)[0]
        ws['B'+str(i+1)]=data["varUnits"][index][0]
    wb.save(Excel_location)

#%% 

# Max_speed=5000
# Speed_Step=500
# Min_speed=0 
# set the current range
# I_max=480
# I_min=1
# I_inc=10
# Op_mode=0 # 0= Motor, 1= Generator , 2 = Motor+Generator
# DC_voltage_list=[400, 600]

concept_ev_exportexcel(mc, work_dir,mot_file_name,Excel_location, Max_speed=5000,Speed_Step=500,Min_speed=0,
                       I_max=480,I_min=1,I_inc=10,Op_mode=2,DC_voltage_list=[100] )
# %%
# concept_ev_exportexcel(mc, work_dir,mot_file_name,Excel_location )

# %%
# 


