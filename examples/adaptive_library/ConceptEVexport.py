#%%
from scipy.io import loadmat
from openpyxl import Workbook
import ansys.motorcad.core as pymotorcad
# define the  mot file 
work_dir="C:\Consusltancy\Working\Concept_ev_maps"
mot_file_name="e9_eMobility_IPM_User"


mot_file=work_dir+"\\"+mot_file_name+".mot"
mc = pymotorcad.MotorCAD(keep_instance_open=True)
mc.load_from_file(mot_file)
mc.set_variable("MessageDisplayState", 2)
mc.set_motorlab_context()

#%%
# set the speed range
Max_speed=5000
Speed_Step=500
Min_speed=0 
mc.set_variable("SpeedMax_MotorLAB",Max_speed)
mc.set_variable("SpeedMin_MotorLAB",Min_speed)
mc.set_variable("Speedinc_MotorLAB",Speed_Step)


# set current range
# check weather a RMS or Peak current def
Current_def=mc.get_variable("CurrentSpec_MotorLAB")
I_max=480
I_min=1
I_step=10
if Current_def==0: # peak 
    mc.set_variable("Imax_MotorLAB",I_max)
    mc.set_variable("Imin_MotorLAB",I_min)
    mc.set_variable("Iinc_MotorLAB",I_step)
else: #RMS
    mc.set_variable("Imax_RMS_MotorLAB",I_max)
    mc.set_variable("Imin_RMS_MotorLAB",I_min)
    mc.set_variable("Iinc_MotorLAB",I_step)

# choose motoring, generating or both modes
Op_mode=0 # 0= Motor, 1= Generator , 2 = Motor+Generator
mc.set_variable("OperatingMode_Lab",Op_mode)
#%%
wb = Workbook()
# choose number of DC bus voltages (list as user input)
#%%
DC_voltage_list=[400, 600]
# ws=wb.create_sheet("Newsheet")
ws=wb.active
ws.title="Voltages"
ws['A1']="Index"
ws['B1']="Voltages"
for i,DC_voltage in enumerate(DC_voltage_list):
    ws['A'+str(i+2)]=i+1
    ws['B'+str(i+2)]=DC_voltage_list[i]
#%%
# set _calcualtion type  Efficiency Map
mc.set_variable("EmagneticCalcType_Lab",1)
sheets=["Speed", "Shaft_Torque","Stator_Current_Line_RMS","Total_Loss","Power_Factor"]
for i,DC_voltage in enumerate(DC_voltage_list):
    mc.set_variable("DCBusVoltage",DC_voltage_list[i])
# run efficiency map calculation 
    mc.calculate_magnetic_lab()
# read the efficieny map .mat file 
    data_file_path=work_dir+"\\"+mot_file_name+"\\Lab\\MotorLAB_elecdata.mat"
    data=loadmat(data_file_path)
    i_len, j_len = data["Speed"].shape
    for sheet in sheets:
        ws=wb.create_sheet("Newsheet")
        ws.title="Voltages"
        if len(DC_voltage_list)>1:
            ws.title=sheet+str(i+1)
        else:
            ws.title=sheet
        ii=0
        jj=0
        for col in ws.iter_cols(min_col=0,max_col=j_len,max_row=i_len):
            ii=0
            for cell in col:
                ws[cell.coordinate]=data[sheet][ii][jj]
                ii=ii+1
            jj=jj+1

Excel_location=work_dir+"\\"+"Map_export.xlsx"
#%%
wb.save(Excel_location)


